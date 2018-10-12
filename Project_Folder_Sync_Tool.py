
from tkinter import *
import time
import openpyxl
from openpyxl import load_workbook
import os,shutil
from gooey import Gooey, GooeyParser
import tkinter.messagebox
import base64
from icon import Icon
from tkinter import filedialog
#import argparse
import distutils.dir_util

import getpass


t0 = time.clock()
running = True

@Gooey(optional_cols=1, program_name="Project Sync and Creator",show_success_modal=0,default_size=(1300, 530))
def main():
    try:
        username=getpass.getuser().upper()
        print("User ID:"+username)
        #print(str.upper(username).replace(".","").replace(" ",""))
        username_formating=str(username).replace(".","").replace(" ","")

        settings_msg = ''

        parser = GooeyParser(description=settings_msg)

        subs = parser.add_subparsers(help='commands', dest='command')

        bom_parser = subs.add_parser('Project_Sync_Creator', help='Project Sync & Creator is a tool to Sync and Create your project')
        bom_parser.add_argument('Target_Folder',
                           help='Choose Target_Folder which contain your projects \n (when Sync option is enabled, this become the source folder \n and target is \\\\tpfs05\\DATA\\RD2\\02_Design Document) and \\\\tpfs05\\DATA\\RD2\\02_Design Document)',
                           type=str, widget='DirChooser')
        #bom_parser.add_argument('--Project_List', type=argparse.FileType,
        #                         help='Write the protocol headers to the specified file')

        bom_parser.add_argument('Project_List',
                           help='Choose HW Project List \n (Please download it from RD2 website)',
                           type=str, widget='FileChooser')


        bom_parser.add_argument("--Sync", action="store_true", help='When Enabled,Target Folder Became Source Folder \n All Folders in Project list will be updated to tpfs05 network drive')

        t_dataproc = time.clock()

        args = parser.parse_args()
        Path = args.Target_Folder
        P_list=args.Project_List
        #print(Path)
        #print(P_list)


        wb = load_workbook(P_list,data_only=True)
        ws = wb.active # Must make it active for the rest of operations

        Base_filename = os.path.basename(P_list)

        if ((ws['A1'].value) != "Project"):
            root = Tk()
            with open('tmp.ico', 'wb') as tmp:
                tmp.write(base64.b64decode(Icon().img))
            root.iconbitmap('tmp.ico')
            os.remove('tmp.ico')
            root.withdraw()
            tkinter.messagebox.showinfo("Format error", "Please make sure Column A1 in '"+ Base_filename+"' is 'Project'")
            wb.close()
            sys.exit()


        row_count = ws.max_row
        #print(row_count)
        i = 2
        if args.Sync is not True:
            while i<=row_count:
                project_rd = ws.cell(i, 12).value
                project_rd_formating = str.upper(project_rd).replace(".", "").replace(" ", "")
                if Path==("\\\\tpfs05\\DATA\\RD2\\02_Design Document") and (project_rd_formating) == (username_formating) :
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\Schematic")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\Layout")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\Test Report")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\Design Package\\EVB Schematic")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\Design Package\\EVB Layout")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\Design Package\\EVB User Guide")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\nvram")
                   print(ws.cell(i,1).value+" project folder created!")


                elif Path==("\\\\tpfs05\\DATA\\RD2\\02_Design Document") or Path==("\\\\tpfs05\\DATA\\RD2\\04_FAE Document") and (project_rd_formating) != (username_formating) :
                   root = Tk()
                   with open('tmp.ico', 'wb') as tmp:
                     tmp.write(base64.b64decode(Icon().img))
                   root.iconbitmap('tmp.ico')
                   os.remove('tmp.ico')
                   root.withdraw()
                   tkinter.messagebox.showinfo("ID Verification Error", "Only '"+ws.cell(i,1).value+"' project owners are allowed for this operation")

                elif Path==("\\\\tpfs05\\DATA\\RD2\\04_FAE Document") and (project_rd_formating) == (username_formating) :

                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\Design Package\\EVB Schematic")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\Design Package\\EVB Layout")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\Design Package\\EVB User Guide")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\nvram")
                   print(ws.cell(i, 1).value + " project folder created!")

                else:
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\Schematic")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\Layout")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\Test Board"+ "\\Layout")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\Test Board" + "\\Schematic")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\Test Board" + "\\For MFG")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\Test Board" + "\\Component Docs")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\For Certification")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\For MFG")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\Component Docs")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\ISO Docs")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\Test Report")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\For Customers"+"\\Design Package\\EVB Schematic")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\For Customers"+"\\Design Package\\EVB Layout")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\For Customers"+"\\Design Package\\EVB User Guide")
                   distutils.dir_util.mkpath(Path + "\\" + ws.cell(i, 1).value + "\\Software"+"\\nvram")
                   print(ws.cell(i, 1).value + " project folder created!")
                i=i+1
        elif args.Sync :
            project_sync(ws,Path,row_count,username_formating)

#        New_Path=os.path.split(Path)[0]+"\\"

#            wb.save(New_Path+Base_filename+"_new.xlsx")
#            wb.close()
#            print("DataPROC=" + str(time.clock() - t_dataproc) + "sec")
#            print("The new xlsx is successfully created")




    except openpyxl.utils.exceptions.InvalidFileException:
             root = Tk()
             with open('tmp.ico', 'wb') as tmp:
                 tmp.write(base64.b64decode(Icon().img))
             root.iconbitmap('tmp.ico')
             os.remove('tmp.ico')
             root.withdraw()
             tkinter.messagebox.showinfo("File format error", "Please use .xlsx file format")

    except FileNotFoundError:
            root = Tk()
            with open('tmp.ico', 'wb') as tmp:
                tmp.write(base64.b64decode(Icon().img))
            root.iconbitmap('tmp.ico')
            os.remove('tmp.ico')
            root.withdraw()
            tkinter.messagebox.showinfo("File format error", "File not found")




def project_sync(ws,Path,row_count,username_formating):

    dest_path1="\\\\tpfs05\\DATA\\RD2\\02_Design Document"
    dest_path2 ="\\\\tpfs05\\DATA\\RD2\\04_FAE Document"

    i = 2
    while i<=row_count:
        project_rd = ws.cell(i, 12).value
        project_rd_formating = str.upper(project_rd).replace(".", "").replace(" ", "")
        #print( project_rd_formating)
        if (project_rd_formating) != (username_formating):
            root = Tk()
            with open('tmp.ico', 'wb') as tmp:
                tmp.write(base64.b64decode(Icon().img))
            root.iconbitmap('tmp.ico')
            os.remove('tmp.ico')
            root.withdraw()
            tkinter.messagebox.showinfo("ID Verification Error","Only '" + ws.cell(i,1).value + "' project owners are allowed for this operation")
        elif (project_rd_formating) == (username_formating):
         try:    ############## For 02_Design Document
                 distutils.dir_util.copy_tree(Path + "\\" + ws.cell(i, 1).value + "\\Schematic",dest_path1+"\\" + ws.cell(i, 1).value + "\\Schematic",update=1)
                 distutils.dir_util.copy_tree(Path + "\\" + ws.cell(i, 1).value + "\\Layout", dest_path1+ "\\" + ws.cell(i, 1).value + "\\Layout", update=1)
                 distutils.dir_util.copy_tree(Path +  "\\" + ws.cell(i, 1).value + "\\Test Report", dest_path1+ "\\" + ws.cell(i, 1).value + "\\Test Report", update=1)
                 distutils.dir_util.copy_tree(Path + "\\" + ws.cell(i, 1).value +  "\\For Customers"+"\\Design Package",dest_path1 + "\\" + ws.cell(i, 1).value + "\\Design Package", update=1)
                 distutils.dir_util.copy_tree(Path + "\\" + ws.cell(i, 1).value +  "\\Software"+"\\nvram",dest_path1 + "\\" + ws.cell(i, 1).value + "\\nvram", update=1)
                 ############## For 04_FAE Document
                 distutils.dir_util.copy_tree(Path +  "\\" + ws.cell(i, 1).value + "\\Test Report", dest_path2+ "\\" + ws.cell(i, 1).value + "\\Test Report", update=1)
                 distutils.dir_util.copy_tree(Path + "\\" + ws.cell(i, 1).value +  "\\For Customers"+"\\Design Package",dest_path2 + "\\" + ws.cell(i, 1).value + "\\Design Package", update=1)
                 distutils.dir_util.copy_tree(Path + "\\" + ws.cell(i, 1).value +  "\\Software"+"\\nvram",dest_path2 + "\\" + ws.cell(i, 1).value + "\\nvram", update=1)
                 print("Project '" + ws.cell(i,1).value + "' Sync Success!")

         except distutils.errors.DistutilsFileError:
                root = Tk()
                with open('tmp.ico', 'wb') as tmp:
                   tmp.write(base64.b64decode(Icon().img))
                root.iconbitmap('tmp.ico')
                os.remove('tmp.ico')
                root.withdraw()
                tkinter.messagebox.showinfo("No Such Folder","Folder missing in project '" + ws.cell(i,1).value + "'")

        i = i + 1
#def handle():

    #print("成功建立'" + Base_filename + "new.xlsx'")
#    print("Done")
#def info_warn_err(Base_filename):
#    root = Tk()
#    with open('tmp.ico', 'wb') as tmp:
#        tmp.write(base64.b64decode(Icon().img))
#    root.iconbitmap('tmp.ico')
#    os.remove('tmp.ico')
#    root.withdraw()
#    tkinter.messagebox.showinfo("Oops!", "Program terminated! '"+Base_filename+"_new.xlsx' is in use")

def version_check():
    try:
        hdk_id=9
        wb = load_workbook("\\\\tpfs05\\DATA\\RD2\\30_Personal Data\\Steven.Jian\\HDK_VER.xlsx")
        ws = wb.active  # Must make it active for the rest of operations
        row_count = ws.max_row
        latest_hdk =ws.cell(row_count, 2).value
        hdk_idnew=int(ws.cell(row_count,1).value)
        wb.close()
        if hdk_idnew is not hdk_id:
            #print(hdk_idnew)
            #print(hdk_id)
            root = Tk()
            with open('tmp.ico', 'wb') as tmp:
                tmp.write(base64.b64decode(Icon().img))
            root.iconbitmap('tmp.ico')
            os.remove('tmp.ico')
            root.withdraw()
            tkinter.messagebox.showinfo("Info","New HDK available! \n Please download it and use the latest bom_tool")
            try:
                dirname = filedialog.asksaveasfilename(filetypes=[("HDK","*.zip")],title="Save the new HDK as...",initialfile=latest_hdk)
                shutil.copy('\\\\tpfs05\\DATA\\RD2\\30_Personal Data\\Steven.Jian\\'+latest_hdk,str(dirname))
                print("HDK download success")
            except OSError:

                root = Tk()
                with open('tmp.ico', 'wb') as tmp:
                    tmp.write(base64.b64decode(Icon().img))
                root.iconbitmap('tmp.ico')
                os.remove('tmp.ico')
                root.withdraw()
                tkinter.messagebox.showinfo("Info", "Download Canceled")
            #win32api.ShellExecute(0, 'open', '\\\\tpfs05\\DATA\\RD2\\30_Personal Data\\Steven.Jian\\'+latest_hdk, '', '', 1)

    except OSError:

        root = Tk()
        with open('tmp.ico', 'wb') as tmp:
            tmp.write(base64.b64decode(Icon().img))
        root.iconbitmap('tmp.ico')
        os.remove('tmp.ico')
        root.withdraw()
        tkinter.messagebox.showinfo("Offline Mode","Please login to azwave.com for HDK version check")

if __name__ == '__main__':
  main()

print("Total="+str(time.clock()-t0)+"sec")
t_vcheck = time.clock()
version_check()
print("version_check()=" + str(time.clock() - t_vcheck) + "sec")
